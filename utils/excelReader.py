import openpyxl
def get_data(path, sheet):
    data_list=[]
    wkbk=openpyxl.load_workbook(path)
    sht=wkbk[sheet]
    tot_r=sht.max_row
    tot_c=sht.max_column
    for r in range(2, tot_r + 1):
        row_list=[]
        for c in range(1, tot_c + 1):
            row_list.append(sht.cell(row=r, column=c).value)
        data_list.append(row_list)
    return data_list
